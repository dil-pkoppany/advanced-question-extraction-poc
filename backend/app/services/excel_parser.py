"""Excel and CSV file parsing utilities."""

import csv
import logging
import re
from pathlib import Path
from typing import Any

import openpyxl
from markitdown import MarkItDown

from ..schemas import SheetMetadata, ColumnMapping, ExtractedQuestion, QuestionType

logger = logging.getLogger(__name__)

# File type detection
EXCEL_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".xltx", ".xltm"}
CSV_EXTENSIONS = {".csv"}


class ExcelParser:
    """Parse Excel files for metadata and content extraction."""

    def __init__(self):
        self.markitdown = MarkItDown()

    def _is_csv(self, file_path: str) -> bool:
        """Check if file is a CSV based on extension."""
        return Path(file_path).suffix.lower() in CSV_EXTENSIONS

    def get_file_metadata(self, file_path: str) -> list[SheetMetadata]:
        """
        Extract metadata from an Excel or CSV file.

        Returns sheet names, column headers, row counts, and sample data.
        """
        if self._is_csv(file_path):
            return self._get_csv_metadata(file_path)
        return self._get_excel_metadata(file_path)

    def _get_csv_metadata(self, file_path: str) -> list[SheetMetadata]:
        """Extract metadata from a CSV file."""
        # CSV files have a single "sheet" named after the file
        file_name = Path(file_path).stem
        
        with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
            # Detect dialect
            sample = f.read(8192)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel
            
            reader = csv.reader(f, dialect)
            rows = list(reader)
        
        if not rows:
            return [SheetMetadata(name=file_name, columns=[], row_count=0, sample_data=[])]
        
        # First row is headers
        columns = []
        for col_idx, value in enumerate(rows[0]):
            if value and value.strip():
                columns.append(value.strip())
            else:
                # Use 0-based column index to match markitdown format
                columns.append(f"Unnamed: {col_idx}")
        
        # Count data rows (excluding header)
        data_rows = rows[1:]
        row_count = sum(1 for row in data_rows if any(cell.strip() for cell in row if cell))
        
        # Get sample data (first 30 rows, first 10 columns, truncate long values)
        MAX_SAMPLE_ROWS = 30
        MAX_SAMPLE_COLS = 10
        MAX_CELL_LENGTH = 200
        sample_data = []
        for row in data_rows[:MAX_SAMPLE_ROWS]:
            row_data = {}
            for col_idx, value in enumerate(row):
                if col_idx >= MAX_SAMPLE_COLS:
                    break  # Only include first 10 columns
                if col_idx < len(columns):
                    if value and value.strip():
                        cell_str = value.strip()
                        if len(cell_str) > MAX_CELL_LENGTH:
                            cell_str = cell_str[:MAX_CELL_LENGTH] + "..."
                        row_data[columns[col_idx]] = cell_str
                    else:
                        row_data[columns[col_idx]] = None
            if any(v is not None for v in row_data.values()):
                sample_data.append(row_data)
        
        return [
            SheetMetadata(
                name=file_name,
                columns=columns,
                row_count=row_count,
                sample_data=sample_data,
            )
        ]

    def _get_excel_metadata(self, file_path: str) -> list[SheetMetadata]:
        """Extract metadata from an Excel file."""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets_metadata = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Get column headers (first row)
            columns = []
            for col_idx, cell in enumerate(ws[1]):
                if cell.value is not None:
                    columns.append(str(cell.value))
                else:
                    # Use 0-based column index to match markitdown format
                    columns.append(f"Unnamed: {col_idx}")

            # Count rows (excluding header)
            row_count = 0
            for row in ws.iter_rows(min_row=2):
                if any(cell.value is not None for cell in row):
                    row_count += 1

            # Get sample data (first 30 rows, first 10 columns, truncate long values)
            MAX_SAMPLE_ROWS = 30
            MAX_SAMPLE_COLS = 10
            MAX_CELL_LENGTH = 200  # Truncate long cell values
            sample_data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=MAX_SAMPLE_ROWS + 1)):
                row_data = {}
                for col_idx, cell in enumerate(row):
                    if col_idx >= MAX_SAMPLE_COLS:
                        break  # Only include first 10 columns
                    if col_idx < len(columns):
                        if cell.value is not None:
                            cell_str = str(cell.value)
                            if len(cell_str) > MAX_CELL_LENGTH:
                                cell_str = cell_str[:MAX_CELL_LENGTH] + "..."
                            row_data[columns[col_idx]] = cell_str
                        else:
                            row_data[columns[col_idx]] = None
                if any(v is not None for v in row_data.values()):
                    sample_data.append(row_data)

            sheets_metadata.append(
                SheetMetadata(
                    name=sheet_name,
                    columns=columns,
                    row_count=row_count,
                    sample_data=sample_data,
                )
            )

        wb.close()
        return sheets_metadata

    def convert_to_markdown(self, file_path: str) -> str:
        """
        Convert Excel or CSV file to Markdown.

        For Excel: Uses MarkItDown
        For CSV: Uses custom conversion to markdown table

        This is used by Approach 1 (Auto LLM).
        """
        if self._is_csv(file_path):
            return self._convert_csv_to_markdown(file_path)
        
        return self._convert_excel_to_markdown(file_path)

    def _convert_csv_to_markdown(self, file_path: str) -> str:
        """Convert CSV file to Markdown table format."""
        with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
            # Detect dialect
            sample = f.read(8192)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel
            
            reader = csv.reader(f, dialect)
            rows = list(reader)
        
        if not rows:
            return ""
        
        # Build markdown table
        lines = []
        file_name = Path(file_path).stem
        lines.append(f"# {file_name}\n")
        
        # Header row
        headers = rows[0]
        lines.append("| " + " | ".join(h if h else "-" for h in headers) + " |")
        lines.append("| " + " | ".join("---" for _ in headers) + " |")
        
        # Data rows
        for row in rows[1:]:
            # Pad row to match header length
            padded_row = row + [""] * (len(headers) - len(row))
            # Clean values
            clean_values = []
            for val in padded_row[:len(headers)]:
                if val and val.strip():
                    # Escape pipe characters and clean whitespace
                    clean_val = val.strip().replace("|", "\\|").replace("\n", " ")
                    clean_values.append(clean_val)
                else:
                    clean_values.append("-")
            lines.append("| " + " | ".join(clean_values) + " |")
        
        markdown_text = "\n".join(lines)
        logger.info(f"CSV converted to {len(markdown_text)} characters of markdown")
        return markdown_text

    def _convert_excel_to_markdown(self, file_path: str) -> str:
        """Convert Excel file to Markdown using MarkItDown."""
        result = self.markitdown.convert(file_path)

        if result and hasattr(result, "text_content"):
            markdown_text = result.text_content

            # Clean up NaN values
            markdown_text = markdown_text.replace(" NaN ", " - ")
            markdown_text = markdown_text.replace("| NaN |", "| - |")
            markdown_text = markdown_text.replace("\nNaN\n", "\n-\n")
            markdown_text = markdown_text.replace("NaN", "-")

            logger.info(f"MarkItDown extracted {len(markdown_text)} characters")
            return markdown_text

        logger.warning("MarkItDown returned empty result")
        return ""

    def generate_filtered_markdown(
        self,
        file_path: str,
        sheet_name: str,
        columns: list[str],
        start_row: int = 1,
        end_row: int | None = None,
        max_rows: int = 100,
        header_row: int = 1,
        column_indices: list[int] | None = None,
    ) -> str:
        """
        Generate markdown table with only specified columns.
        
        Used by Approach 4 to pass filtered data to Step 3.
        
        Args:
            file_path: Path to Excel/CSV file
            sheet_name: Name of sheet to extract from
            columns: List of column names to include (e.g., ["Unnamed: 4", "Unnamed: 5"])
            start_row: Row number to start extraction (1-based, default 1)
            end_row: Row number to end extraction (1-based, None for all rows)
            max_rows: Maximum number of rows to include (for token control)
            header_row: Row number containing headers (1-based, default 1)
            column_indices: Optional pre-resolved column indices (0-based). When provided,
                these take priority over name-based column lookup for Excel files.
                This handles the mismatch between pandas column names (from Step 1)
                and openpyxl header row column names (used in Step 3).
            
        Returns:
            Markdown table string with filtered columns
        """
        if self._is_csv(file_path):
            return self._generate_filtered_markdown_csv(
                file_path, sheet_name, columns, start_row, end_row, max_rows
            )
        return self._generate_filtered_markdown_excel(
            file_path, sheet_name, columns, start_row, end_row, max_rows, header_row,
            column_indices,
        )

    def _generate_filtered_markdown_csv(
        self,
        file_path: str,
        sheet_name: str,
        columns: list[str],
        start_row: int,
        end_row: int | None,
        max_rows: int,
    ) -> str:
        """Generate filtered markdown from CSV file."""
        file_name = Path(file_path).stem
        
        # CSV files have a single "sheet" named after the file
        if sheet_name != file_name:
            logger.warning(f"Sheet '{sheet_name}' not found in CSV (has '{file_name}')")
            return ""
        
        with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
            sample = f.read(8192)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel
            
            reader = csv.reader(f, dialect)
            rows = list(reader)
        
        if not rows:
            return ""
        
        # Build column name mapping (first row is header)
        all_headers = []
        for col_idx, value in enumerate(rows[0]):
            if value and value.strip():
                all_headers.append(value.strip())
            else:
                all_headers.append(f"Unnamed: {col_idx}")
        
        # Find indices of requested columns
        col_indices = []
        filtered_headers = []
        for col_name in columns:
            if col_name in all_headers:
                idx = all_headers.index(col_name)
                col_indices.append(idx)
                filtered_headers.append(col_name)
            else:
                # Fallback: handle "Unnamed: N" style columns by positional index
                unnamed_match = re.match(r"^Unnamed:\s*(\d+)$", col_name)
                if unnamed_match:
                    positional_idx = int(unnamed_match.group(1))
                    if positional_idx < len(all_headers):
                        col_indices.append(positional_idx)
                        filtered_headers.append(all_headers[positional_idx])
                        logger.info(
                            f"CSV column '{col_name}' resolved by positional index "
                            f"{positional_idx} -> '{all_headers[positional_idx]}'"
                        )
                else:
                    # Try case-insensitive match
                    col_name_lower = col_name.lower().strip()
                    for h_idx, h in enumerate(all_headers):
                        if h.lower().strip() == col_name_lower:
                            col_indices.append(h_idx)
                            filtered_headers.append(all_headers[h_idx])
                            break
        
        if not col_indices:
            logger.warning(f"No matching columns found in CSV for: {columns}")
            return ""
        
        # Build markdown table
        lines = []
        
        # Header row
        lines.append("| Row | " + " | ".join(filtered_headers) + " |")
        lines.append("| --- | " + " | ".join("---" for _ in filtered_headers) + " |")
        
        # Data rows
        data_rows = rows[1:]
        actual_start = max(0, start_row - 2)  # Convert to 0-based index
        actual_end = min(len(data_rows), (end_row - 1) if end_row else len(data_rows))
        
        rows_added = 0
        for data_idx in range(actual_start, actual_end):
            if rows_added >= max_rows:
                break
            
            row = data_rows[data_idx]
            row_num = data_idx + 2  # 1-based row number (accounting for header)
            
            # Extract values for filtered columns
            values = []
            for col_idx in col_indices:
                if col_idx < len(row):
                    val = row[col_idx]
                    if val and val.strip():
                        # Escape pipe characters and clean whitespace
                        clean_val = val.strip().replace("|", "\\|").replace("\n", " ")
                        values.append(clean_val)
                    else:
                        values.append("-")
                else:
                    values.append("-")
            
            lines.append(f"| {row_num} | " + " | ".join(values) + " |")
            rows_added += 1
        
        markdown_text = "\n".join(lines)
        logger.info(f"Generated filtered markdown: {len(filtered_headers)} columns, {rows_added} rows")
        return markdown_text

    def _generate_filtered_markdown_excel(
        self,
        file_path: str,
        sheet_name: str,
        columns: list[str],
        start_row: int,
        end_row: int | None,
        max_rows: int,
        header_row: int,
        column_indices: list[int] | None = None,
    ) -> str:
        """Generate filtered markdown from Excel file."""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            wb.close()
            return ""
        
        ws = wb[sheet_name]
        
        # Get column headers from specified header row
        all_headers = []
        header_row_data = list(ws[header_row])
        for col_idx, cell in enumerate(header_row_data):
            if cell.value is not None:
                all_headers.append(str(cell.value))
            else:
                all_headers.append(f"Unnamed: {col_idx}")
        
        # Resolve column indices.
        # Priority: use pre-resolved indices from caller (avoids pandas/openpyxl name mismatch),
        # then fall back to name-based lookup.
        col_indices: list[int] = []
        filtered_headers: list[str] = []
        
        if column_indices:
            # Use pre-resolved indices directly (from pandas metadata)
            seen = set()
            for idx in column_indices:
                if idx < len(all_headers) and idx not in seen:
                    col_indices.append(idx)
                    filtered_headers.append(all_headers[idx])
                    seen.add(idx)
                elif idx >= len(all_headers):
                    logger.warning(
                        f"Pre-resolved column index {idx} out of range "
                        f"(sheet '{sheet_name}' has {len(all_headers)} columns in header row {header_row})"
                    )
            if col_indices:
                logger.info(
                    f"Using {len(col_indices)} pre-resolved column indices for sheet '{sheet_name}': "
                    f"{list(zip(col_indices, filtered_headers))}"
                )
        
        # Fall back to name-based lookup if no pre-resolved indices or none matched
        if not col_indices:
            for col_name in columns:
                if col_name in all_headers:
                    # Direct match found
                    idx = all_headers.index(col_name)
                    col_indices.append(idx)
                    filtered_headers.append(all_headers[idx])
                else:
                    # Fallback: handle "Unnamed: N" style columns from pandas
                    unnamed_match = re.match(r"^Unnamed:\s*(\d+)$", col_name)
                    if unnamed_match:
                        positional_idx = int(unnamed_match.group(1))
                        if positional_idx < len(all_headers):
                            col_indices.append(positional_idx)
                            filtered_headers.append(all_headers[positional_idx])
                            logger.info(
                                f"Column '{col_name}' not found in header row {header_row} of sheet '{sheet_name}', "
                                f"using positional index {positional_idx} -> '{all_headers[positional_idx]}'"
                            )
                        else:
                            logger.warning(
                                f"Column '{col_name}' positional index {positional_idx} out of range "
                                f"(sheet '{sheet_name}' has {len(all_headers)} columns)"
                            )
                    else:
                        # Try case-insensitive match as a last resort
                        col_name_lower = col_name.lower().strip()
                        matched = False
                        for h_idx, h in enumerate(all_headers):
                            if h.lower().strip() == col_name_lower:
                                col_indices.append(h_idx)
                                filtered_headers.append(all_headers[h_idx])
                                logger.info(
                                    f"Column '{col_name}' matched case-insensitively to '{all_headers[h_idx]}' "
                                    f"in sheet '{sheet_name}'"
                                )
                                matched = True
                                break
                        if not matched:
                            logger.warning(
                                f"Column '{col_name}' not found in sheet '{sheet_name}' "
                                f"header row {header_row}. Available: {all_headers}"
                            )
        
        if not col_indices:
            logger.warning(f"No matching columns found in sheet '{sheet_name}' for: {columns}")
            wb.close()
            return ""
        
        # Build markdown table
        lines = []
        
        # Header row (use actual column names from Excel)
        display_headers = []
        for col_idx in col_indices:
            header_text = all_headers[col_idx]
            # Escape pipe characters
            header_text = header_text.replace("|", "\\|").replace("\n", " ")
            display_headers.append(header_text)
        
        lines.append("| Row | " + " | ".join(display_headers) + " |")
        lines.append("| --- | " + " | ".join("---" for _ in display_headers) + " |")
        
        # Data rows
        actual_start = max(start_row, header_row + 1)  # Start after header
        actual_end = end_row if end_row else ws.max_row
        
        rows_added = 0
        for row_idx in range(actual_start, actual_end + 1):
            if rows_added >= max_rows:
                break
            
            row = list(ws[row_idx])
            
            # Extract values for filtered columns
            values = []
            has_content = False
            for col_idx in col_indices:
                if col_idx < len(row):
                    cell = row[col_idx]
                    if cell.value is not None:
                        val = str(cell.value).strip()
                        if val:
                            has_content = True
                            # Escape pipe characters and clean whitespace
                            clean_val = val.replace("|", "\\|").replace("\n", " ")
                            # Truncate very long values
                            if len(clean_val) > 500:
                                clean_val = clean_val[:500] + "..."
                            values.append(clean_val)
                        else:
                            values.append("-")
                    else:
                        values.append("-")
                else:
                    values.append("-")
            
            # Only include rows that have some content
            if has_content:
                lines.append(f"| {row_idx} | " + " | ".join(values) + " |")
                rows_added += 1
        
        wb.close()
        
        markdown_text = "\n".join(lines)
        logger.info(f"Generated filtered markdown for '{sheet_name}': {len(filtered_headers)} columns, {rows_added} rows")
        return markdown_text

    def extract_rows_by_columns(
        self,
        file_path: str,
        column_mappings: list[ColumnMapping],
    ) -> list[ExtractedQuestion]:
        """
        Extract questions deterministically from specified columns.

        This is used by Approaches 2 and 3.
        Returns raw extracted data without LLM processing.
        """
        if self._is_csv(file_path):
            return self._extract_rows_from_csv(file_path, column_mappings)
        return self._extract_rows_from_excel(file_path, column_mappings)

    def _extract_rows_from_csv(
        self,
        file_path: str,
        column_mappings: list[ColumnMapping],
    ) -> list[ExtractedQuestion]:
        """Extract questions from CSV file."""
        file_name = Path(file_path).stem
        questions = []
        
        with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
            sample = f.read(8192)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel
            
            reader = csv.reader(f, dialect)
            rows = list(reader)
        
        if not rows:
            return questions
        
        # Build column name mapping (same logic as _get_csv_metadata)
        headers = []
        for col_idx, value in enumerate(rows[0]):
            if value and value.strip():
                headers.append(value.strip())
            else:
                headers.append(f"Column_{col_idx + 1}")
        
        for mapping in column_mappings:
            # CSV files have a single "sheet" named after the file
            if mapping.sheet_name != file_name:
                logger.warning(f"Sheet '{mapping.sheet_name}' not found (CSV has sheet '{file_name}')")
                continue
            
            # Find column indices
            col_indices = {}
            for idx, col_name in enumerate(headers):
                if col_name == mapping.question_column:
                    col_indices["question"] = idx
                if mapping.answer_column and col_name == mapping.answer_column:
                    col_indices["answer"] = idx
                if mapping.type_column and col_name == mapping.type_column:
                    col_indices["type"] = idx
            
            if "question" not in col_indices:
                logger.warning(
                    f"Question column '{mapping.question_column}' not found in '{mapping.sheet_name}'"
                )
                continue
            
            # Extract rows (row indices are 1-based, header is row 1, data starts at row 2)
            data_rows = rows[1:]
            start_idx = mapping.start_row - 2  # Convert to 0-based index into data_rows
            end_idx = (mapping.end_row - 1) if mapping.end_row else len(data_rows)
            
            for data_idx in range(max(0, start_idx), min(end_idx, len(data_rows))):
                row = data_rows[data_idx]
                row_idx = data_idx + 2  # Convert back to 1-based row number
                
                # Get question text
                q_val = row[col_indices["question"]] if col_indices["question"] < len(row) else ""
                question_text = q_val.strip() if q_val else ""
                
                if not question_text or question_text == "-":
                    continue
                
                # Get answers if available
                answers = None
                if "answer" in col_indices and col_indices["answer"] < len(row):
                    a_val = row[col_indices["answer"]]
                    if a_val and a_val.strip():
                        answer_text = a_val
                        if "|" in answer_text:
                            answers = [a.strip() for a in answer_text.split("|")]
                        elif "\n" in answer_text:
                            answers = [a.strip() for a in answer_text.split("\n")]
                        else:
                            answers = [answer_text.strip()]
                
                # Determine question type
                question_type = QuestionType.OPEN_ENDED
                if "type" in col_indices and col_indices["type"] < len(row):
                    t_val = row[col_indices["type"]]
                    if t_val and t_val.strip():
                        type_str = t_val.lower().strip()
                        type_mapping = {
                            "open": QuestionType.OPEN_ENDED,
                            "open_ended": QuestionType.OPEN_ENDED,
                            "single": QuestionType.SINGLE_CHOICE,
                            "single_choice": QuestionType.SINGLE_CHOICE,
                            "multiple": QuestionType.MULTIPLE_CHOICE,
                            "multiple_choice": QuestionType.MULTIPLE_CHOICE,
                            "grouped": QuestionType.GROUPED_QUESTION,
                            "grouped_question": QuestionType.GROUPED_QUESTION,
                            "yes_no": QuestionType.YES_NO,
                            "yesno": QuestionType.YES_NO,
                        }
                        question_type = type_mapping.get(type_str, QuestionType.OPEN_ENDED)
                elif answers:
                    if len(answers) == 2 and set(a.lower() for a in answers) <= {"yes", "no"}:
                        question_type = QuestionType.YES_NO
                    else:
                        question_type = QuestionType.SINGLE_CHOICE
                
                questions.append(
                    ExtractedQuestion(
                        question_text=question_text,
                        question_type=question_type,
                        answers=answers,
                        row_index=row_idx,
                        sheet_name=mapping.sheet_name,
                    )
                )
        
        logger.info(f"Extracted {len(questions)} questions from CSV columns")
        return questions

    def _extract_rows_from_excel(
        self,
        file_path: str,
        column_mappings: list[ColumnMapping],
    ) -> list[ExtractedQuestion]:
        """Extract questions from Excel file."""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        questions = []

        for mapping in column_mappings:
            if mapping.sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{mapping.sheet_name}' not found")
                continue

            ws = wb[mapping.sheet_name]

            # Find column indices - use same naming as get_file_metadata
            header_row = list(ws[1])
            col_indices = {}
            for idx, cell in enumerate(header_row):
                # Generate column name same way as get_file_metadata
                if cell.value is not None:
                    col_name = str(cell.value)
                else:
                    # Use 0-based column index to match markitdown format
                    col_name = f"Unnamed: {idx}"
                
                if col_name == mapping.question_column:
                    col_indices["question"] = idx
                if mapping.answer_column and col_name == mapping.answer_column:
                    col_indices["answer"] = idx
                if mapping.type_column and col_name == mapping.type_column:
                    col_indices["type"] = idx

            if "question" not in col_indices:
                logger.warning(
                    f"Question column '{mapping.question_column}' not found in '{mapping.sheet_name}'"
                )
                continue

            # Extract rows
            start_row = mapping.start_row
            end_row = mapping.end_row or ws.max_row

            for row_idx in range(start_row, end_row + 1):
                row = list(ws[row_idx])

                # Get question text
                q_cell = row[col_indices["question"]] if col_indices["question"] < len(row) else None
                question_text = str(q_cell.value).strip() if q_cell and q_cell.value else ""

                if not question_text or question_text == "-":
                    continue

                # Get answers if available
                answers = None
                if "answer" in col_indices:
                    a_cell = row[col_indices["answer"]] if col_indices["answer"] < len(row) else None
                    if a_cell and a_cell.value:
                        answer_text = str(a_cell.value)
                        # Parse pipe-separated or newline-separated answers
                        if "|" in answer_text:
                            answers = [a.strip() for a in answer_text.split("|")]
                        elif "\n" in answer_text:
                            answers = [a.strip() for a in answer_text.split("\n")]
                        else:
                            answers = [answer_text.strip()]

                # Determine question type
                question_type = QuestionType.OPEN_ENDED
                if "type" in col_indices:
                    t_cell = row[col_indices["type"]] if col_indices["type"] < len(row) else None
                    if t_cell and t_cell.value:
                        type_str = str(t_cell.value).lower().strip()
                        type_mapping = {
                            "open": QuestionType.OPEN_ENDED,
                            "open_ended": QuestionType.OPEN_ENDED,
                            "single": QuestionType.SINGLE_CHOICE,
                            "single_choice": QuestionType.SINGLE_CHOICE,
                            "multiple": QuestionType.MULTIPLE_CHOICE,
                            "multiple_choice": QuestionType.MULTIPLE_CHOICE,
                            "grouped": QuestionType.GROUPED_QUESTION,
                            "grouped_question": QuestionType.GROUPED_QUESTION,
                            "yes_no": QuestionType.YES_NO,
                            "yesno": QuestionType.YES_NO,
                        }
                        question_type = type_mapping.get(type_str, QuestionType.OPEN_ENDED)
                elif answers:
                    # Infer type from answers
                    if len(answers) == 2 and set(a.lower() for a in answers) <= {"yes", "no"}:
                        question_type = QuestionType.YES_NO
                    else:
                        question_type = QuestionType.SINGLE_CHOICE

                questions.append(
                    ExtractedQuestion(
                        question_text=question_text,
                        question_type=question_type,
                        answers=answers,
                        row_index=row_idx,
                        sheet_name=mapping.sheet_name,
                    )
                )

        wb.close()
        logger.info(f"Extracted {len(questions)} questions from columns")
        return questions

    def count_rows_in_columns(
        self,
        file_path: str,
        column_mappings: list[ColumnMapping],
    ) -> int:
        """
        Count non-empty rows in the specified question columns.

        Used for validation in Approaches 2 and 3.
        """
        if self._is_csv(file_path):
            return self._count_rows_in_csv(file_path, column_mappings)
        return self._count_rows_in_excel(file_path, column_mappings)

    def _count_rows_in_csv(
        self,
        file_path: str,
        column_mappings: list[ColumnMapping],
    ) -> int:
        """Count non-empty rows in CSV file."""
        file_name = Path(file_path).stem
        total_count = 0
        
        with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
            sample = f.read(8192)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel
            
            reader = csv.reader(f, dialect)
            rows = list(reader)
        
        if not rows:
            return 0
        
        # Build column name mapping
        headers = []
        for col_idx, value in enumerate(rows[0]):
            if value and value.strip():
                headers.append(value.strip())
            else:
                headers.append(f"Column_{col_idx + 1}")
        
        for mapping in column_mappings:
            if mapping.sheet_name != file_name:
                continue
            
            # Find question column index
            q_col_idx = None
            for idx, col_name in enumerate(headers):
                if col_name == mapping.question_column:
                    q_col_idx = idx
                    break
            
            if q_col_idx is None:
                continue
            
            # Count non-empty rows
            data_rows = rows[1:]
            start_idx = mapping.start_row - 2
            end_idx = (mapping.end_row - 1) if mapping.end_row else len(data_rows)
            
            for data_idx in range(max(0, start_idx), min(end_idx, len(data_rows))):
                row = data_rows[data_idx]
                if q_col_idx < len(row):
                    val = row[q_col_idx]
                    if val and val.strip() and val.strip() != "-":
                        total_count += 1
        
        return total_count

    def _count_rows_in_excel(
        self,
        file_path: str,
        column_mappings: list[ColumnMapping],
    ) -> int:
        """Count non-empty rows in Excel file."""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        total_count = 0

        for mapping in column_mappings:
            if mapping.sheet_name not in wb.sheetnames:
                continue

            ws = wb[mapping.sheet_name]

            # Find question column index - use same naming as get_file_metadata
            header_row = list(ws[1])
            q_col_idx = None
            for idx, cell in enumerate(header_row):
                if cell.value is not None:
                    col_name = str(cell.value)
                else:
                    col_name = f"Column_{idx + 1}"
                if col_name == mapping.question_column:
                    q_col_idx = idx
                    break

            if q_col_idx is None:
                continue

            # Count non-empty rows
            start_row = mapping.start_row
            end_row = mapping.end_row or ws.max_row

            for row_idx in range(start_row, end_row + 1):
                row = list(ws[row_idx])
                if q_col_idx < len(row):
                    cell = row[q_col_idx]
                    if cell.value and str(cell.value).strip() and str(cell.value).strip() != "-":
                        total_count += 1

        wb.close()
        return total_count
