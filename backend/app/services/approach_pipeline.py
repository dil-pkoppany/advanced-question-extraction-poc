"""Approach 4: Multi-Step Pipeline Extraction.

This approach uses a multi-step pipeline:
1. Structure Analysis: Identify question/answer columns
2. Coverage Validation: Judge validates structure completeness
3. Question Extraction: Extract questions with types, answers, dependencies
4. Normalization: Convert to Python objects

All intermediate results are saved for POC debugging.
"""

import json
import logging
import time
from pathlib import Path
from typing import Any

import boto3
from botocore.config import Config
from bs4 import BeautifulSoup

from ..config import get_settings
from ..schemas import (
    ExtractedQuestion,
    ExtractionResult,
    ExtractionMetrics,
    QuestionType,
    QuestionDependency,
    SheetMetadata,
)
from .excel_parser import ExcelParser

# Sonnet 4.5 token limits for structure analysis chunking
SONNET_MAX_INPUT_TOKENS = 180000  # Reserve 20K for output from 200K context
CHARS_PER_TOKEN = 4  # Conservative estimate
MAX_INPUT_CHARS = SONNET_MAX_INPUT_TOKENS * CHARS_PER_TOKEN  # 720K chars

# Per-sheet extraction limits
MAX_SAMPLE_COLS = 10
MAX_SAMPLE_ROWS = 30

# Step 3 extraction limits (for filtered markdown)
MAX_ROWS_PER_BATCH = 100  # Maximum rows per LLM batch to control token usage

logger = logging.getLogger(__name__)


class PipelineExtractionService:
    """Multi-step pipeline extraction with intermediate result saving."""

    def __init__(self, model_id: str | None = None):
        self.settings = get_settings()
        self.parser = ExcelParser()
        self.model_id = model_id or self.settings.bedrock_model_id
        # Use Sonnet 4.5 for judge step (Step 2: Coverage Validation)
        self.judge_model_id = self.settings.bedrock_sonnet_model_id

        # Initialize Bedrock clients
        config = Config(
            read_timeout=600,
            connect_timeout=60,
            retries={"max_attempts": 3, "mode": "adaptive"},
        )
        self.bedrock = boto3.client(
            "bedrock-runtime",
            region_name=self.settings.aws_region,
            config=config,
        )

        # Track metrics
        self.total_llm_calls = 0
        self.structure_analysis_time_ms = 0
        self.coverage_validation_time_ms = 0
        self.extraction_time_ms = 0
        self.normalization_time_ms = 0

    def _estimate_sheet_chars(self, sheet: SheetMetadata) -> int:
        """Estimate character count for a sheet's prompt contribution.
        
        Uses first 10 columns and 30 rows of sample data for estimation.
        """
        estimated = len(f"\nSheet: {sheet.name}\n")
        estimated += len(f"Columns: {', '.join(sheet.columns[:MAX_SAMPLE_COLS])}\n")
        estimated += len(f"Row count: {sheet.row_count}\n")
        estimated += len("Sample rows:\n")
        
        for idx, row in enumerate(sheet.sample_data[:MAX_SAMPLE_ROWS]):
            # Only include first 10 columns in the row data
            limited_row = {k: v for i, (k, v) in enumerate(row.items()) if i < MAX_SAMPLE_COLS}
            estimated += len(f"  Row {idx + 2}: {json.dumps(limited_row, default=str)}\n")
        
        return estimated

    def _chunk_sheets_for_structure_analysis(
        self,
        metadata: list[SheetMetadata],
        max_chars: int = MAX_INPUT_CHARS
    ) -> list[list[SheetMetadata]]:
        """Group sheets into chunks that fit within token limit.
        
        Rules:
        - Never split a sheet in half
        - Estimate size based on columns + 30 rows of sample data
        - If a single sheet exceeds limit, it goes alone (may need truncation)
        - Greedily add sheets until approaching limit
        
        Args:
            metadata: List of sheet metadata from Excel file
            max_chars: Maximum characters per chunk (default: Sonnet 4.5 safe limit)
            
        Returns:
            List of chunks, each chunk is a list of SheetMetadata objects
        """
        if not metadata:
            return []
        
        # Reserve space for prompt template (instructions, XML format example)
        prompt_overhead = 1500  # Approximate chars for prompt instructions
        available_chars = max_chars - prompt_overhead
        
        chunks: list[list[SheetMetadata]] = []
        current_chunk: list[SheetMetadata] = []
        current_chunk_size = 0
        
        for sheet in metadata:
            sheet_size = self._estimate_sheet_chars(sheet)
            
            # If adding this sheet would exceed limit, start new chunk
            if current_chunk_size + sheet_size > available_chars:
                if current_chunk:
                    chunks.append(current_chunk)
                # Start new chunk with this sheet
                current_chunk = [sheet]
                current_chunk_size = sheet_size
            else:
                # Add to current chunk
                current_chunk.append(sheet)
                current_chunk_size += sheet_size
        
        # Don't forget the last chunk
        if current_chunk:
            chunks.append(current_chunk)
        
        logger.info(f"Chunked {len(metadata)} sheets into {len(chunks)} chunk(s) for structure analysis")
        return chunks

    def _build_structure_prompt(self, sheets: list[SheetMetadata]) -> str:
        """Build the structure analysis prompt for a chunk of sheets.
        
        Args:
            sheets: List of SheetMetadata objects to include in the prompt
            
        Returns:
            Complete prompt string for structure analysis
        """
        prompt_parts = [
            "Analyze this Excel file structure and identify for EACH sheet:",
            "1. Which columns contain questions (question_column)",
            "2. Which columns contain answer values/checkboxes (answer_column) - often TRUE/FALSE or checkbox values",
            "3. Which columns contain answer option TEXT/labels (answer_options_column) - the human-readable text for each answer choice",
            "4. Header row location",
            "5. Data start row (where actual questions begin - may be row 30+ in some files)",
            "",
            "IMPORTANT: answer_column and answer_options_column may be DIFFERENT columns:",
            "- answer_column: Contains the actual answer values (TRUE/FALSE, checkbox states, selected values)",
            "- answer_options_column: Contains the TEXT labels describing each answer option",
            "For checkbox-based surveys, the checkbox alt-text/labels are often in a separate column from the checkbox values.",
            "",
            "Sheets:",
        ]
        
        for sheet in sheets:
            prompt_parts.append(f"\nSheet: {sheet.name}")
            prompt_parts.append(f"Columns: {', '.join(sheet.columns[:MAX_SAMPLE_COLS])}")
            prompt_parts.append(f"Row count: {sheet.row_count}")
            if sheet.sample_data:
                prompt_parts.append(f"Sample rows (first {min(len(sheet.sample_data), MAX_SAMPLE_ROWS)}):")
                for idx, row in enumerate(sheet.sample_data[:MAX_SAMPLE_ROWS]):
                    # Only include first 10 columns in the row data
                    limited_row = {k: v for i, (k, v) in enumerate(row.items()) if i < MAX_SAMPLE_COLS}
                    prompt_parts.append(f"  Row {idx + 2}: {json.dumps(limited_row, default=str)}")
        
        prompt_parts.extend([
            "",
            "Respond in XML format:",
            "<structure_analysis>",
            '  <sheet sheet_name="SheetName" header_row="1" data_start_row="2" confidence="0.95">',
            '    <columns question_column="Column_5" answer_column="Column_6" answer_options_column="Column_7" type_column="" instruction_column=""/>',
            '    <structure_notes>Questions in column 5, checkbox values in column 6, answer option text/labels in column 7</structure_notes>',
            "  </sheet>",
            "</structure_analysis>",
            "",
            "IMPORTANT: Include a <sheet> element for EACH sheet in the input.",
            "IMPORTANT: answer_options_column should contain the TEXT labels for answer choices, which may be different from answer_column (checkbox values).",
            "Return ONLY the XML."
        ])
        
        return "\n".join(prompt_parts)

    async def extract(self, file_path: str, run_id: str | None = None) -> ExtractionResult:
        """
        Extract questions using multi-step pipeline.

        Args:
            file_path: Path to the Excel file
            run_id: Optional run ID for saving intermediate results

        Returns:
            ExtractionResult with questions and metrics
        """
        start_time = time.time()
        run_id = run_id or f"pipeline_{int(time.time())}"
        intermediate_dir = None

        try:
            # Create intermediate results directory
            intermediate_dir = None
            if run_id:
                run_dir = self.settings.runs_dir / run_id
                run_dir.mkdir(parents=True, exist_ok=True)
                intermediate_dir = run_dir / "intermediate_results"
                intermediate_dir.mkdir(parents=True, exist_ok=True)

            # Step 1: Structure Analysis
            logger.info("Step 1: Analyzing structure...")
            step1_start = time.time()
            structure = await self._analyze_structure(file_path, intermediate_dir)
            self.structure_analysis_time_ms = int((time.time() - step1_start) * 1000)
            
            if intermediate_dir:
                self._save_intermediate_result(
                    intermediate_dir, 1, structure, run_id, "structure_analysis.json"
                )

            if not structure or not structure.get("sheets"):
                return ExtractionResult(
                    approach=4,
                    success=False,
                    error="Failed to analyze structure",
                )

            # Step 2: Coverage Validation
            logger.info("Step 2: Validating coverage...")
            step2_start = time.time()
            coverage = await self._validate_coverage(structure, file_path, intermediate_dir)
            self.coverage_validation_time_ms = int((time.time() - step2_start) * 1000)
            
            if intermediate_dir:
                self._save_intermediate_result(
                    intermediate_dir, 2, coverage, run_id, "coverage_validation.json"
                )

            # Step 3: Question Extraction
            logger.info("Step 3: Extracting questions...")
            step3_start = time.time()
            xml_output = await self._extract_questions(structure, file_path, intermediate_dir, run_id)
            self.extraction_time_ms = int((time.time() - step3_start) * 1000)

            if not xml_output:
                return ExtractionResult(
                    approach=4,
                    success=False,
                    error="Failed to extract questions",
                )

            # Step 4: Normalization
            logger.info("Step 4: Normalizing to objects...")
            step4_start = time.time()
            questions = self._normalize_to_objects(xml_output)
            self.normalization_time_ms = int((time.time() - step4_start) * 1000)
            
            if intermediate_dir:
                self._save_intermediate_result(
                    intermediate_dir, 4, [q.model_dump() for q in questions], run_id, "normalized_questions.json"
                )

            # Calculate metrics
            total_time_ms = int((time.time() - start_time) * 1000)
            show_deps_count = sum(
                1 for q in questions
                if q.dependencies and any(d.dependency_action == "show" for d in q.dependencies)
            )
            skip_deps_count = sum(
                1 for q in questions
                if q.dependencies and any(d.dependency_action == "skip" for d in q.dependencies)
            )

            return ExtractionResult(
                approach=4,
                success=True,
                questions=questions,
                metrics=ExtractionMetrics(
                    extraction_count=len(questions),
                    total_time_ms=total_time_ms,
                    structure_analysis_time_ms=self.structure_analysis_time_ms,
                    coverage_validation_time_ms=self.coverage_validation_time_ms,
                    extraction_time_ms=self.extraction_time_ms,
                    normalization_time_ms=self.normalization_time_ms,
                    total_llm_calls=self.total_llm_calls,
                    structure_confidence=structure.get("confidence"),
                    coverage_confidence=coverage.get("confidence"),
                    show_dependencies_count=show_deps_count,
                    skip_dependencies_count=skip_deps_count,
                ),
            )

        except Exception as e:
            logger.error(f"Pipeline extraction failed: {e}", exc_info=True)
            return ExtractionResult(
                approach=4,
                success=False,
                error=str(e),
                metrics=ExtractionMetrics(
                    extraction_count=0,
                    total_time_ms=int((time.time() - start_time) * 1000),
                ),
            )

    async def _analyze_structure(self, file_path: str, intermediate_dir: Path | None = None) -> dict[str, Any]:
        """Step 1: Analyze Excel structure to identify question/answer columns.
        
        Uses chunked processing to handle large files with many sheets.
        Each chunk contains complete sheets (never split) that fit within
        Sonnet 4.5's token limit.
        """
        # Get metadata for all sheets
        metadata = self.parser.get_file_metadata(file_path)
        
        if not metadata:
            logger.error("No sheets found in file")
            return {"sheets": [], "confidence": 0.0}
        
        # Chunk sheets by token limit
        chunks = self._chunk_sheets_for_structure_analysis(metadata)
        
        if not chunks:
            logger.error("Failed to chunk sheets")
            return {"sheets": [], "confidence": 0.0}
        
        all_sheets: list[dict] = []
        total_confidence = 0.0
        chunk_count = len(chunks)
        
        logger.info(f"Processing {len(metadata)} sheets in {chunk_count} chunk(s)")
        
        for chunk_idx, sheet_chunk in enumerate(chunks):
            chunk_num = chunk_idx + 1
            sheet_names = [s.name for s in sheet_chunk]
            logger.info(f"Processing chunk {chunk_num}/{chunk_count}: sheets {sheet_names}")
            
            # Build prompt for this chunk's sheets
            prompt = self._build_structure_prompt(sheet_chunk)
            
            # Save prompt
            if intermediate_dir:
                if chunk_count == 1:
                    # Single chunk - use original filename for backward compatibility
                    prompt_file = intermediate_dir / "step1_structure_analysis_prompt.txt"
                else:
                    prompt_file = intermediate_dir / f"step1_structure_chunk_{chunk_num}_prompt.txt"
                prompt_file.write_text(prompt)
            
            # Call LLM
            try:
                response = await self._invoke_llm(prompt, self.model_id, "xml")
                self.total_llm_calls += 1
            except Exception as e:
                logger.error(f"LLM call failed for chunk {chunk_num}: {e}")
                continue
            
            # Save response
            if intermediate_dir:
                if chunk_count == 1:
                    response_file = intermediate_dir / "step1_structure_analysis_response.xml"
                else:
                    response_file = intermediate_dir / f"step1_structure_chunk_{chunk_num}_response.xml"
                response_file.write_text(response)
            
            if not response or not response.strip():
                logger.warning(f"Empty response from chunk {chunk_num}")
                continue
            
            # Parse and merge results
            try:
                result = self._parse_structure_xml(response)
                chunk_sheets = result.get("sheets", [])
                chunk_confidence = result.get("confidence", 0.0)
                
                all_sheets.extend(chunk_sheets)
                total_confidence += chunk_confidence
                
                logger.info(f"Chunk {chunk_num}: found {len(chunk_sheets)} sheet structure(s), confidence={chunk_confidence}")
            except Exception as e:
                logger.error(f"Failed to parse structure analysis for chunk {chunk_num}: {e}")
                logger.debug(f"Response was: {response[:500]}")
        
        # Calculate average confidence
        avg_confidence = total_confidence / chunk_count if chunk_count > 0 else 0.0
        
        merged_result = {
            "sheets": all_sheets,
            "confidence": avg_confidence,
        }
        
        logger.info(f"Structure analysis complete: {len(all_sheets)} sheets, avg confidence={avg_confidence:.2f}")
        
        return merged_result

    async def _validate_coverage(self, structure: dict, file_path: str, intermediate_dir: Path | None = None) -> dict[str, Any]:
        """Step 2: Validate structure analysis completeness."""
        metadata = self.parser.get_file_metadata(file_path)
        
        prompt = f"""Validate this structure analysis. Check:
1. Are all question columns identified?
2. Are answer options properly located?
3. Are there missing elements?

Structure Analysis:
{json.dumps(structure, indent=2)}

Available Sheets:
{json.dumps([{"name": s.name, "columns": s.columns, "row_count": s.row_count} for s in metadata], indent=2)}

Respond in XML format:
<coverage_validation is_complete="true" confidence="0.92">
  <missing_elements/>
  <suggestions>
    <suggestion>Column_7 may contain conditional logic indicators</suggestion>
  </suggestions>
</coverage_validation>

Return ONLY the XML."""
        
        # Save prompt
        if intermediate_dir:
            prompt_file = intermediate_dir / "step2_coverage_validation_prompt.txt"
            prompt_file.write_text(prompt)
        
        response = await self._invoke_llm(prompt, self.judge_model_id, "xml")
        self.total_llm_calls += 1
        
        # Save response
        if intermediate_dir:
            response_file = intermediate_dir / "step2_coverage_validation_response.xml"
            response_file.write_text(response)
        
        if not response or not response.strip():
            logger.warning("Empty response from coverage validation")
            return {"is_complete": True, "missing_elements": [], "suggestions": [], "confidence": 0.5}
        
        try:
            return self._parse_coverage_xml(response)
        except Exception as e:
            logger.error(f"Failed to parse coverage validation: {e}")
            logger.debug(f"Response was: {response[:500]}")
            return {"is_complete": True, "missing_elements": [], "suggestions": [], "confidence": 0.5}

    async def _extract_questions(
        self, structure: dict, file_path: str, intermediate_dir: Path | None, run_id: str
    ) -> str:
        """Step 3: Extract questions with types, answers, dependencies."""
        import re
        
        # Extract filtered markdown for each sheet
        batches = self._extract_with_context(structure, file_path)
        
        all_xml_parts = []
        batch_num = 0
        
        for batch in batches:
            batch_num += 1
            
            # Get the actual sheet name from the batch dict
            actual_sheet_name = batch.get("sheet_name", "Sheet1")
            
            # Build prompt with filtered markdown
            prompt = self._build_extraction_prompt(batch)
            
            # Save prompt
            if intermediate_dir:
                prompt_file = intermediate_dir / f"step3_extraction_batch_{batch_num}_prompt.txt"
                prompt_file.write_text(prompt)
            
            response = await self._invoke_llm(prompt, self.model_id, "xml")
            self.total_llm_calls += 1
            
            # Save raw response before any modification
            if intermediate_dir:
                raw_response_file = intermediate_dir / f"step3_extraction_batch_{batch_num}_response_raw.xml"
                raw_response_file.write_text(response)
            
            # Replace any sheet name in the XML with the actual sheet name
            # The LLM might output sheet="Sheet1" but we know the real sheet name
            response = re.sub(
                r'sheet="[^"]*"',
                f'sheet="{actual_sheet_name}"',
                response
            )
            
            # Save batch XML (with corrected sheet name)
            if intermediate_dir:
                xml_file = intermediate_dir / f"step3_question_extraction_batch_{batch_num}.xml"
                xml_file.write_text(response)
            
            all_xml_parts.append(response)
        
        # Combine all batches
        combined_xml = "<questions>\n" + "\n".join(
            part.replace("<questions>", "").replace("</questions>", "").strip()
            for part in all_xml_parts
        ) + "\n</questions>"
        
        if intermediate_dir:
            combined_file = intermediate_dir / "step3_question_extraction_combined.xml"
            combined_file.write_text(combined_xml)
        
        return combined_xml

    def _extract_with_context(self, structure: dict, file_path: str) -> list[dict]:
        """Extract filtered markdown data for each sheet.
        
        Uses sheet-based batching with row limits: generates filtered markdown per sheet,
        splitting large sheets into multiple batches.
        
        The markdown table preserves visual structure, making it easier for the LLM to:
        - See multi-row answer options as table rows
        - Understand column relationships from headers
        - Identify question types from patterns
        
        Returns list of dicts with {sheet_name, markdown, header_row, data_start_row, columns, batch_num}
        """
        import openpyxl
        
        batches = []
        
        # Get row counts for each sheet to determine batching
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_row_counts = {}
        for sheet_name in wb.sheetnames:
            sheet_row_counts[sheet_name] = wb[sheet_name].max_row
        wb.close()
        
        # Build pandas-style column name -> index mapping from file metadata.
        # Step 1 uses pandas which reads row 1 as the header, producing names like
        # "02.01 Contact Information" or "Unnamed: 3". But Step 3's markdown generator
        # uses openpyxl with the actual header_row (e.g., row 4), where column names
        # may differ ("Question", "Response", etc.). We need to resolve pandas names
        # to positional indices so the markdown generator can find the right columns.
        metadata = self.parser.get_file_metadata(file_path)
        pandas_col_to_index: dict[str, dict[str, int]] = {}  # sheet_name -> {col_name -> index}
        if metadata:
            for sheet_meta in metadata:
                col_map = {}
                for idx, col_name in enumerate(sheet_meta.columns):
                    col_map[col_name] = idx
                pandas_col_to_index[sheet_meta.name] = col_map
        
        for sheet_info in structure.get("sheets", []):
            sheet_name = sheet_info["sheet_name"]
            columns_info = sheet_info.get("columns", {})
            header_row = sheet_info.get("header_row", 1)
            data_start = sheet_info.get("data_start_row", 2)
            
            # Get question column - required
            question_col = columns_info.get("question_column")
            if not question_col:
                logger.warning(f"No question column identified for sheet '{sheet_name}', skipping")
                continue
            
            # Build list of columns to include in filtered markdown
            # Start with question column, then add other relevant columns
            columns_to_include = [question_col]
            
            # Add answer column if present
            if columns_info.get("answer_column"):
                columns_to_include.append(columns_info["answer_column"])
            
            # Add answer options column if present (important for checkbox labels)
            if columns_info.get("answer_options_column"):
                columns_to_include.append(columns_info["answer_options_column"])
            
            # Add other useful columns
            for col_key in ["type_column", "instruction_column", "additional_answer_column", "followup_column"]:
                if columns_info.get(col_key):
                    col_name = columns_info[col_key]
                    if col_name not in columns_to_include:
                        columns_to_include.append(col_name)
            
            # Resolve column names to positional indices using pandas metadata.
            # This handles the mismatch where pandas names differ from openpyxl header names.
            # The LLM may alter column names (e.g., dropping '&' or special chars),
            # so we use normalized matching as a fallback.
            col_indices_to_include = []
            sheet_col_map = pandas_col_to_index.get(sheet_name, {})
            
            # Build a normalized lookup for fuzzy matching
            def _normalize_col_name(name: str) -> str:
                """Normalize column name for fuzzy matching: lowercase, collapse whitespace, strip special chars."""
                import re as _re
                normalized = name.lower().strip()
                # Remove common special chars the LLM might drop
                normalized = _re.sub(r'[&@#$%^*()+=\[\]{}<>|\\/:;\'\"~`]', ' ', normalized)
                # Collapse multiple whitespace to single space
                normalized = _re.sub(r'\s+', ' ', normalized).strip()
                return normalized
            
            normalized_col_map: dict[str, int] = {}
            for meta_col_name, meta_idx in sheet_col_map.items():
                normalized_col_map[_normalize_col_name(meta_col_name)] = meta_idx
            
            for col_name in columns_to_include:
                # Try exact match first
                if col_name in sheet_col_map:
                    idx = sheet_col_map[col_name]
                    if idx not in col_indices_to_include:
                        col_indices_to_include.append(idx)
                else:
                    # Try normalized/fuzzy match (handles LLM dropping '&' etc.)
                    norm_name = _normalize_col_name(col_name)
                    if norm_name in normalized_col_map:
                        idx = normalized_col_map[norm_name]
                        if idx not in col_indices_to_include:
                            col_indices_to_include.append(idx)
                            logger.info(
                                f"Column '{col_name}' matched via normalized lookup to index {idx} "
                                f"in sheet '{sheet_name}'"
                            )
                    else:
                        logger.warning(
                            f"Column '{col_name}' not found in pandas metadata for sheet '{sheet_name}', "
                            f"will attempt name-based lookup in header row"
                        )
            
            # Get total row count for this sheet
            total_rows = sheet_row_counts.get(sheet_name, 0)
            data_rows = total_rows - data_start + 1  # Rows from data_start to end
            
            # Calculate number of batches needed
            num_batches = max(1, (data_rows + MAX_ROWS_PER_BATCH - 1) // MAX_ROWS_PER_BATCH)
            
            # Generate batches for this sheet
            sheet_batch_num = 0
            current_start = data_start
            
            while current_start <= total_rows:
                sheet_batch_num += 1
                
                # Generate filtered markdown for this batch
                # Pass both column names and resolved indices for robust lookup
                markdown = self.parser.generate_filtered_markdown(
                    file_path=file_path,
                    sheet_name=sheet_name,
                    columns=columns_to_include,
                    start_row=current_start,
                    end_row=None,  # Will be limited by max_rows
                    max_rows=MAX_ROWS_PER_BATCH,
                    header_row=header_row,
                    column_indices=col_indices_to_include if col_indices_to_include else None,
                )
                
                if not markdown:
                    # No more data rows
                    break
                
                # Extract actual column headers from the markdown table.
                # The prompt should reference the actual table headers (from the openpyxl
                # header row), not the pandas names which may differ.
                resolved_question_col = columns_info.get("question_column", "Question")
                resolved_answer_col = columns_info.get("answer_column", "")
                resolved_answer_options_col = columns_info.get("answer_options_column", "")
                
                # Parse actual headers from the first line of the markdown table
                md_lines = markdown.strip().split("\n")
                if md_lines:
                    header_line = md_lines[0]
                    # Parse "| Row | Header1 | Header2 | ..." format
                    md_headers = [h.strip() for h in header_line.split("|") if h.strip()]
                    # Skip "Row" column
                    actual_headers = [h for h in md_headers if h != "Row"]
                    
                    if actual_headers:
                        # Map resolved indices to actual header names.
                        # columns_to_include[0] is always question_column
                        if len(actual_headers) >= 1:
                            resolved_question_col = actual_headers[0]
                        if len(actual_headers) >= 2:
                            if columns_info.get("answer_options_column"):
                                resolved_answer_options_col = actual_headers[1]
                            else:
                                resolved_answer_col = actual_headers[1]
                        if len(actual_headers) >= 3:
                            # Third column is answer_options if separate, or additional column
                            if columns_info.get("answer_options_column") and columns_info.get("answer_column"):
                                resolved_answer_col = actual_headers[2]
                
                # Build resolved_columns_info with actual header names
                resolved_columns_info = {
                    "question_column": resolved_question_col,
                    "answer_column": resolved_answer_col,
                    "answer_options_column": resolved_answer_options_col,
                }
                
                # Create batch info
                batch = {
                    "sheet_name": sheet_name,
                    "markdown": markdown,
                    "header_row": header_row,
                    "data_start_row": current_start,
                    "columns": columns_to_include,
                    "columns_info": columns_info,
                    "resolved_columns_info": resolved_columns_info,
                    "batch_num": sheet_batch_num,
                    "total_batches": num_batches,
                }
                
                batches.append(batch)
                
                # Move to next batch
                current_start += MAX_ROWS_PER_BATCH
                
                # Safety check - don't create too many batches
                if sheet_batch_num >= 20:  # Max 20 batches per sheet
                    logger.warning(f"Sheet '{sheet_name}' has too many rows, limiting to {sheet_batch_num} batches")
                    break
        
        return batches if batches else []

    def _build_extraction_prompt(self, batch: dict) -> str:
        """Build extraction prompt for a batch using filtered markdown table.
        
        The batch contains:
        - sheet_name: Name of the sheet
        - markdown: Filtered markdown table with relevant columns
        - columns: List of column names included
        - columns_info: Original column mapping from Step 1
        - resolved_columns_info: Column names as they appear in the markdown table headers
        """
        sheet_name = batch.get("sheet_name", "Sheet1")
        markdown = batch.get("markdown", "")
        
        # Use resolved column names (actual table headers) when available,
        # fall back to original columns_info (pandas names) for backwards compatibility
        resolved = batch.get("resolved_columns_info", {})
        columns_info = batch.get("columns_info", {})
        
        # Get column names for context - prefer resolved names that match the table
        question_col = resolved.get("question_column") or columns_info.get("question_column", "Question")
        answer_col = resolved.get("answer_column") or columns_info.get("answer_column", "")
        answer_options_col = resolved.get("answer_options_column") or columns_info.get("answer_options_column", "")
        
        prompt_parts = [
            f"Extract questions from this Excel sheet data (sheet: '{sheet_name}').",
            "",
            "The data is provided as a markdown table. Each row has a Row number in the first column.",
            "",
            "IMPORTANT - Multi-row questions:",
            "- Rows with '-' in the question column but content in answer columns are CONTINUATION rows",
            "- These continuation rows contain additional answer OPTIONS for the previous question",
            "- Group all continuation rows with their parent question",
            "",
            "For each question:",
            "1. Separate question text from instructions/comments",
            "2. Identify type (open_ended, single_choice, multiple_choice, numeric, yes_no, etc.)",
            "3. Extract ALL answer options from the question row AND its continuation rows",
            "4. Parse dependencies:",
            "   - Show: 'appears only if...' → action='show'",
            "   - Skip: 'skip if...', 'hidden when...' → action='skip'",
            "5. Detect FOLLOW-UP questions (indicate conditional dependencies):",
            "   - Text patterns: 'If you can not...', 'If no...', 'If not...', 'Please explain...'",
            "   - When detected, create dependency to PREVIOUS question row number",
            "   - Action is 'show' (follow-up appears when main question answered negatively)",
            "",
            "Question type guidelines:",
            "- yes_no: EXACTLY 2 options that are simple 'Yes'/'No' or 'True'/'False' binary choices",
            "- single_choice: Multiple options but only one can be selected",
            "- multiple_choice: Multiple options with checkboxes (multiple can be selected)",
            "  HINT: If you see TRUE/FALSE values in one column and text labels in another column,",
            "  the text labels are the answer OPTIONS and TRUE/FALSE indicates selected state",
            "- open_ended: No predefined answer options, free text input",
            "",
        ]
        
        # Add column context if available
        if answer_options_col:
            prompt_parts.extend([
                f"Column info: Question text is in '{question_col}', answer option TEXT is in '{answer_options_col}'.",
                f"The '{answer_options_col}' column contains the human-readable labels for each answer option.",
                "",
            ])
        elif answer_col:
            prompt_parts.extend([
                f"Column info: Question text is in '{question_col}', answers are in '{answer_col}'.",
                "",
            ])
        
        # Add the markdown table
        prompt_parts.extend([
            "DATA:",
            markdown,
            "",
        ])
        
        prompt_parts.extend([
            "",
            "Output XML format:",
            '<questions>',
            '  <q type="single_choice" row="2" sheet="Sheet1">',
            '    <text>What is your industry?</text>',
            '    <help_text>Please select one</help_text>',
            '    <answers><option>Manufacturing</option><option>Services</option></answers>',
            '    <dependencies><depends_on question_row="5" answer_value="Yes" action="show"/></dependencies>',
            '  </q>',
            '  <q type="yes_no" row="5" sheet="Sheet1">',
            '    <text>Do you have international operations?</text>',
            '    <help_text></help_text>',
            '    <answers><option>Yes</option><option>No</option></answers>',
            '    <dependencies></dependencies>',
            '  </q>',
            '  <q type="yes_no" row="8" sheet="Sheet1">',
            '    <text>Do you have sustainability certifications?</text>',
            '    <help_text></help_text>',
            '    <answers><option>Yes</option><option>No</option></answers>',
            '    <conditional_inputs><input answer="Yes">please provide detail about which certifications</input></conditional_inputs>',
            '    <dependencies></dependencies>',
            '  </q>',
            '  <q type="single_choice" row="21" sheet="Sheet1">',
            '    <text>Has your company been audited by an independent external auditor?</text>',
            '    <help_text>If yes, upload the audit report</help_text>',
            '    <answers>',
            '      <option>Yes, virtual audit</option>',
            '      <option>Yes, on-site audit</option>',
            '      <option>No audit yet/I don\'t know</option>',
            '    </answers>',
            '    <dependencies></dependencies>',
            '  </q>',
            '  <q type="multiple_choice" row="15" sheet="Sheet1">',
            '    <text>Does your company have any of the following certifications?</text>',
            '    <help_text></help_text>',
            '    <answers>',
            '      <option>Environmental certifications, such as ISO 50001, ISO 14001, EMAS</option>',
            '      <option>Labor and human rights certifications, such as Fair Wage Network</option>',
            '      <option>Business ethics certification(s), such as ISO 27001</option>',
            '    </answers>',
            '    <dependencies></dependencies>',
            '  </q>',
            '  <!-- Follow-up question example: main question and dependent follow-up -->',
            '  <q type="yes_no" row="30" sheet="Sheet1">',
            '    <text>Can you confirm you meet this requirement?</text>',
            '    <help_text></help_text>',
            '    <answers><option>Yes</option><option>No</option></answers>',
            '    <dependencies></dependencies>',
            '  </q>',
            '  <q type="open_ended" row="31" sheet="Sheet1">',
            '    <text>If you can not reach this requirement, please detail here.</text>',
            '    <help_text></help_text>',
            '    <answers></answers>',
            '    <dependencies>',
            '      <depends_on question_row="30" answer_value="No" action="show"/>',
            '    </dependencies>',
            '  </q>',
            '</questions>',
            "",
            "Return ONLY the XML."
        ])
        
        return "\n".join(prompt_parts)

    def _normalize_to_objects(self, xml: str) -> list[ExtractedQuestion]:
        """Step 4: Convert XML to ExtractedQuestion objects.
        
        This method generates unique GUIDs for each question and resolves
        dependency references from sheet:row format to GUIDs.
        """
        import uuid
        
        questions = []
        # Map sheet:row -> question_id (GUID) for dependency resolution
        location_to_guid: dict[str, str] = {}
        # Store raw dependency data for second pass
        raw_dependencies: dict[int, list[dict]] = {}
        
        try:
            soup = BeautifulSoup(xml, "xml")
            questions_tag = soup.find("questions")
            
            if not questions_tag:
                return questions
            
            # First pass: create questions with GUIDs and build location map
            for idx, q_tag in enumerate(questions_tag.find_all("q")):
                question_text = q_tag.find("text")
                if not question_text:
                    continue
                
                q_text = question_text.get_text(strip=True)
                q_type_str = q_tag.get("type", "open_ended")
                
                # Map type
                type_mapping = {
                    "open_ended": QuestionType.OPEN_ENDED,
                    "single_choice": QuestionType.SINGLE_CHOICE,
                    "multiple_choice": QuestionType.MULTIPLE_CHOICE,
                    "grouped_question": QuestionType.GROUPED_QUESTION,
                    "yes_no": QuestionType.YES_NO,
                    "numeric": QuestionType.NUMERIC,
                    "integer": QuestionType.INTEGER,
                    "decimal": QuestionType.DECIMAL,
                }
                question_type = type_mapping.get(q_type_str, QuestionType.OPEN_ENDED)
                
                # Help text
                help_text_tag = q_tag.find("help_text")
                help_text = help_text_tag.get_text(strip=True) if help_text_tag else None
                if help_text == "":
                    help_text = None
                
                # Answers
                answers = None
                answers_tag = q_tag.find("answers")
                if answers_tag:
                    options = answers_tag.find_all("option")
                    if options:
                        answers = [opt.get_text(strip=True) for opt in options]
                
                # Conditional inputs (for answers that require additional detail)
                conditional_inputs = None
                cond_inputs_tag = q_tag.find("conditional_inputs")
                if cond_inputs_tag:
                    input_tags = cond_inputs_tag.find_all("input")
                    if input_tags:
                        conditional_inputs = {}
                        for input_tag in input_tags:
                            answer_value = input_tag.get("answer", "").strip()
                            input_prompt = input_tag.get_text(strip=True)
                            if answer_value and input_prompt:
                                conditional_inputs[answer_value] = input_prompt
                        if not conditional_inputs:
                            conditional_inputs = None
                
                # Generate GUID for this question
                question_id = str(uuid.uuid4())
                
                # Store location -> GUID mapping
                question_sheet = q_tag.get("sheet", "")
                row_index = int(q_tag.get("row")) if q_tag.get("row") else None
                if question_sheet and row_index is not None:
                    location_key = f"{question_sheet}:{row_index}"
                    location_to_guid[location_key] = question_id
                
                # Store raw dependencies for second pass
                deps_tag = q_tag.find("dependencies")
                if deps_tag:
                    raw_deps = []
                    for dep_tag in deps_tag.find_all("depends_on"):
                        raw_dep_id = dep_tag.get("question_row") or dep_tag.get("question_id") or ""
                        raw_deps.append({
                            "raw_id": raw_dep_id,
                            "sheet": question_sheet,
                            "answer_value": dep_tag.get("answer_value"),
                            "condition_type": dep_tag.get("condition_type", "equals"),
                            "action": dep_tag.get("action", "show"),
                            "original_text": dep_tag.get("original_text"),
                        })
                    if raw_deps:
                        raw_dependencies[len(questions)] = raw_deps
                
                questions.append(
                    ExtractedQuestion(
                        question_id=question_id,
                        question_text=q_text,
                        question_type=question_type,
                        answers=answers,
                        help_text=help_text,
                        conditional_inputs=conditional_inputs,
                        dependencies=None,  # Will be set in second pass
                        row_index=row_index,
                        sheet_name=question_sheet if question_sheet else None,
                    )
                )
            
            # Second pass: resolve dependencies using GUIDs
            for q_idx, raw_deps in raw_dependencies.items():
                dependencies = []
                for raw_dep in raw_deps:
                    # Build location key to lookup the GUID
                    raw_id = raw_dep["raw_id"]
                    sheet = raw_dep["sheet"]
                    
                    # Create composite key for lookup
                    if sheet and raw_id:
                        location_key = f"{sheet}:{raw_id}"
                    else:
                        location_key = raw_id
                    
                    # Resolve to GUID if found, otherwise keep the raw reference
                    resolved_id = location_to_guid.get(location_key, location_key)
                    
                    dep = QuestionDependency(
                        depends_on_question_id=resolved_id,
                        depends_on_answer_value=raw_dep["answer_value"],
                        condition_type=raw_dep["condition_type"],
                        dependency_action=raw_dep["action"],
                        original_text=raw_dep["original_text"],
                    )
                    dependencies.append(dep)
                
                if dependencies:
                    questions[q_idx].dependencies = dependencies
        
        except Exception as e:
            logger.error(f"XML normalization error: {e}", exc_info=True)
        
        return questions

    async def _invoke_llm(self, prompt: str, model_id: str, response_format: str) -> str:
        """Invoke Bedrock LLM."""
        # Determine max_tokens based on model
        # Opus 4.5 supports 32K, Sonnet 4.5 supports 16K
        if "opus" in model_id.lower():
            max_tokens = 32768  # Opus 4.5 maximum
        elif "sonnet" in model_id.lower():
            max_tokens = 16384  # Sonnet 4.5 maximum
        else:
            # Fallback to settings
            max_tokens = self.settings.max_tokens if response_format == "xml" else self.settings.judge_max_tokens
        
        # Use lower temperature for JSON responses (judge steps)
        temperature = self.settings.temperature if response_format == "xml" else self.settings.judge_temperature
        
        payload = {
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": max_tokens,
            "temperature": temperature,
            "messages": [{"role": "user", "content": prompt}],
        }
        
        response = self.bedrock.invoke_model(
            modelId=model_id,
            body=json.dumps(payload),
            contentType="application/json",
        )
        
        response_body = json.loads(response["body"].read())
        
        if "content" in response_body and len(response_body["content"]) > 0:
            return response_body["content"][0]["text"]
        
        raise Exception("No content in Bedrock response")

    def _parse_structure_xml(self, xml: str) -> dict[str, Any]:
        """Parse structure analysis XML response."""
        try:
            # Find XML content
            xml_start = xml.find("<structure_analysis>")
            xml_end = xml.rfind("</structure_analysis>")
            
            if xml_start < 0:
                return {"sheets": [], "confidence": 0.0}
            
            if xml_end < 0:
                xml_text = xml[xml_start:] + "</structure_analysis>"
            else:
                xml_text = xml[xml_start : xml_end + len("</structure_analysis>")]
            
            soup = BeautifulSoup(xml_text, "xml")
            structure_tag = soup.find("structure_analysis")
            
            if not structure_tag:
                return {"sheets": [], "confidence": 0.0}
            
            sheets = []
            sheet_confidences = []
            for sheet_tag in structure_tag.find_all("sheet"):
                columns_tag = sheet_tag.find("columns")
                columns = {}
                if columns_tag:
                    columns = {
                        "question_column": columns_tag.get("question_column", ""),
                        "answer_column": columns_tag.get("answer_column") or None,
                        "answer_options_column": columns_tag.get("answer_options_column") or None,
                        "type_column": columns_tag.get("type_column") or None,
                        "instruction_column": columns_tag.get("instruction_column") or None,
                        "additional_answer_column": columns_tag.get("additional_answer_column") or None,
                        "followup_column": columns_tag.get("followup_column") or None,
                    }
                
                structure_notes_tag = sheet_tag.find("structure_notes")
                structure_notes = structure_notes_tag.get_text(strip=True) if structure_notes_tag else ""
                
                # Per-sheet confidence
                sheet_confidence = float(sheet_tag.get("confidence", 0.0)) if sheet_tag.get("confidence") else 0.0
                sheet_confidences.append(sheet_confidence)
                
                sheets.append({
                    "sheet_name": sheet_tag.get("sheet_name", ""),
                    "header_row": int(sheet_tag.get("header_row", 1)),
                    "data_start_row": int(sheet_tag.get("data_start_row", 2)),
                    "columns": columns,
                    "structure_notes": structure_notes,
                })
            
            # Confidence: prefer root-level attribute, fall back to average of per-sheet confidences
            confidence = float(structure_tag.get("confidence", 0.0)) if structure_tag.get("confidence") else 0.0
            if confidence == 0.0 and sheet_confidences:
                confidence = sum(sheet_confidences) / len(sheet_confidences)
            
            return {
                "sheets": sheets,
                "confidence": confidence,
            }
        except Exception as e:
            logger.error(f"XML parsing error in structure analysis: {e}")
            return {"sheets": [], "confidence": 0.0}

    def _parse_coverage_xml(self, xml: str) -> dict[str, Any]:
        """Parse coverage validation XML response."""
        try:
            # Find XML content
            xml_start = xml.find("<coverage_validation>")
            xml_end = xml.rfind("</coverage_validation>")
            
            if xml_start < 0:
                return {"is_complete": True, "missing_elements": [], "suggestions": [], "confidence": 0.5}
            
            if xml_end < 0:
                xml_text = xml[xml_start:] + "</coverage_validation>"
            else:
                xml_text = xml[xml_start : xml_end + len("</coverage_validation>")]
            
            soup = BeautifulSoup(xml_text, "xml")
            coverage_tag = soup.find("coverage_validation")
            
            if not coverage_tag:
                return {"is_complete": True, "missing_elements": [], "suggestions": [], "confidence": 0.5}
            
            is_complete = coverage_tag.get("is_complete", "true").lower() == "true"
            confidence = float(coverage_tag.get("confidence", 0.5)) if coverage_tag.get("confidence") else 0.5
            
            missing_elements = []
            missing_tag = coverage_tag.find("missing_elements")
            if missing_tag:
                for elem in missing_tag.find_all("element"):
                    missing_elements.append(elem.get_text(strip=True))
            
            suggestions = []
            suggestions_tag = coverage_tag.find("suggestions")
            if suggestions_tag:
                for suggestion in suggestions_tag.find_all("suggestion"):
                    suggestions.append(suggestion.get_text(strip=True))
            
            return {
                "is_complete": is_complete,
                "missing_elements": missing_elements,
                "suggestions": suggestions,
                "confidence": confidence,
            }
        except Exception as e:
            logger.error(f"XML parsing error in coverage validation: {e}")
            return {"is_complete": True, "missing_elements": [], "suggestions": [], "confidence": 0.5}

    def _save_intermediate_result(
        self, intermediate_dir: Path, step: int, data: dict | list | str, run_id: str, filename: str
    ) -> None:
        """Save intermediate result to file."""
        try:
            result_file = intermediate_dir / filename
            if isinstance(data, str):
                result_file.write_text(data)
            else:
                result_file.write_text(json.dumps(data, indent=2, default=str))
            logger.info(f"Saved intermediate result: {filename}")
        except Exception as e:
            logger.error(f"Failed to save intermediate result {filename}: {e}")
